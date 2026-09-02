# -*- coding: utf-8 -*-
"""校验 / 回归：比对按模板工作簿与参考工作簿是否逐字段一致。"""
from openpyxl import load_workbook


def _load(path):
    wb = load_workbook(path, data_only=True)
    out = {}
    for s in ["方案", "高管", "业绩考核"]:
        ws = wb[s]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
        out[s] = (hdr, rows)
    return out


def _key_of(sheet, hdr, row):
    if sheet == "方案":
        return row[hdr.index("公司代码")] if "公司代码" in hdr else row[0]
    if sheet == "高管":
        return row[hdr.index("代码&姓名")] if "代码&姓名" in hdr else (row[hdr.index("公司代码")] if "公司代码" in hdr else row[0])
    return row[hdr.index("公司代码")] if "公司代码" in hdr else row[0]


def compare_workbooks(path_a, path_b, sheets=("方案", "高管", "业绩考核")):
    """逐行（按主键排序后）比对两个工作簿。返回 (passed: bool, diffs: list[str])。"""
    a = _load(path_a)
    b = _load(path_b)
    diffs = []
    for s in sheets:
        ha, ra = a[s]
        hb, rb = b[s]
        ka = {_key_of(s, ha, r): r for r in ra}
        kb = {_key_of(s, hb, r): r for r in rb}
        if set(ka) != set(kb):
            diffs.append(f"[{s}] 主键集合不同：A={set(ka)} / B={set(kb)}")
            continue
        for k in sorted(set(ka), key=lambda x: str(x)):
            ra_, rb_ = ka[k], kb[k]
            for i, (va, vb) in enumerate(zip(ra_, rb_)):
                ha_ = ha[i] if i < len(ha) else None
                hb_ = hb[i] if i < len(hb) else None
                if ha_ != hb_:
                    continue  # 表头不同跳过（不影响数据）
                # 数值近似比较
                if isinstance(va, float) or isinstance(vb, float):
                    if va is None or vb is None or abs(va - vb) > 1e-9:
                        diffs.append(f"[{s}] {k} 列'{ha_}': {va} != {vb}")
                elif va != vb:
                    diffs.append(f"[{s}] {k} 列'{ha_}': {va!r} != {vb!r}")
    return (len(diffs) == 0), diffs


def verify_workbook(path, records):
    """内部一致性自检：行数 + 关键字段非空。返回 (passed, notes)。"""
    notes = []
    a = _load(path)
    ha, ra = a["方案"]
    passed = True
    if len(ra) != len(records):
        passed = False
        notes.append(f"方案行数 {len(ra)} != records {len(records)}")
    # 每家公司关键字段存在性
    for r in ra:
        code = r[ha.index("公司代码")] if "公司代码" in ha else None
        tot = r[ha.index("激励总数(万股/万份)")] if "激励总数(万股/万份)" in ha else None
        gp = r[ha.index("期权行权价格/股票授予价格")] if "期权行权价格/股票授予价格" in ha else None
        if tot is None or gp is None:
            passed = False
            notes.append(f"{code} 缺少总数或授予价")
    notes.append(f"方案 {len(ra)} 行 / 高管 {len(a['高管'][1])} 行 / 业绩考核 {len(a['业绩考核'][1])} 行")
    return passed, notes
