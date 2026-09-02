# -*- coding: utf-8 -*-
"""
可选：把 store 中的 records 同步进用户桌面的 6 表累积库（全部A股/1-方案/2-授予/3-高管/4-业绩考核/5-公允价值）。
- 仅 1-方案 / 3-高管 / 4-业绩考核 与模板列结构对齐，按列名解析写入；其余表（2-授予/5-公允价值/全部A股）需单独映射，本版跳过并提示。
- 备份优先；按公告ID+公司代码+预案公告日去重；小批追加到 max_row+1。
- 累积库文件当前不在 Desktop（见前期确认）→ 未配置路径时直接返回提示，不做任何改动。
"""
import os
import shutil
from openpyxl import load_workbook
from lti_insight.core.fill_excel import make_builders
from lti_insight.config.loader import load_settings

PCT = "0.00%"
SHEET_MAP = {"方案": "1-方案", "高管": "3-高管", "业绩考核": "4-业绩考核"}
_PCT_HINT = ("比例(%)", "占比", "折扣")


def _resolver(hdr):
    def get(row, name, default=None):
        if name in hdr:
            return row[hdr.index(name)]
        for i, h in enumerate(hdr):
            if isinstance(h, str) and h.startswith(name):
                return row[i]
        return default
    return get


def sync_to_cumulative(records, target_path=None, backup=True):
    cfg = load_settings()
    target_path = target_path or cfg["paths"].get("cumulative_db")
    if not target_path:
        return {"ok": False, "msg": "未配置累积库路径（settings.paths.cumulative_db 为空），跳过同步。"}
    if not os.path.exists(target_path):
        return {"ok": False, "msg": f"累积库文件不存在：{target_path}。请提供真实 6 表累积库路径后再同步。"}

    if backup:
        bak = target_path + ".bak"
        shutil.copy2(target_path, bak)

    wb = load_workbook(target_path)
    report = {"ok": True, "added": {}, "skipped_sheets": []}

    for logical, tab in SHEET_MAP.items():
        if tab not in wb.sheetnames:
            report["skipped_sheets"].append(f"{tab}（不存在）")
            continue
        ws = wb[tab]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        HEAD = {logical: hdr}
        scheme_row, exec_rows_all, perf_row = make_builders(HEAD)
        get = _resolver(hdr)

        # 既有行去重键
        code_idx = hdr.index("公司代码") if "公司代码" in hdr else None
        date_idx = hdr.index("预案公告日") if "预案公告日" in hdr else None
        name_idx = hdr.index("方案名称") if "方案名称" in hdr else None
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if code_idx is None or date_idx is None:
                break
            key = (row[code_idx], row[date_idx], row[name_idx] if name_idx is not None else None)
            existing.add(key)

        if logical == "方案":
            new_rows = [scheme_row(r) for r in records]
        elif logical == "高管":
            new_rows = exec_rows_all(records)
        else:
            new_rows = [perf_row(r) for r in records]

        # 逐条追加并去重
        start = ws.max_row + 1
        if logical == "高管":
            for er in new_rows:
                key = (get(er, "公司代码"), get(er, "预案公告日"), get(er, "方案名称"))
                if key in existing:
                    continue
                _write_row(ws, start, hdr, er)
                existing.add(key)
                start += 1
                added += 1
        else:
            for rr in new_rows:
                key = (get(rr, "公司代码"), get(rr, "预案公告日"), get(rr, "方案名称"))
                if key in existing:
                    continue
                _write_row(ws, start, hdr, rr)
                existing.add(key)
                start += 1
                added += 1
        report["added"][tab] = added

    wb.save(target_path)
    report["backup"] = target_path + ".bak"
    return report


def _write_row(ws, row_idx, hdr, row):
    for c, v in enumerate(row, 1):
        cell = ws.cell(row_idx, c, v)
        h = hdr[c - 1]
        if isinstance(h, str) and any(k in h for k in _PCT_HINT) and isinstance(v, (int, float)):
            cell.number_format = PCT
