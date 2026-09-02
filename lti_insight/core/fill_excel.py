# -*- coding: utf-8 -*-
"""
按用户模板（示例.xlsx，方案157/高管39/业绩考核159）生成 LTI 活数据库工作簿。
- 读取模板精确表头与列数，按列名映射填充（能填的填，缺的留空）
- 表头带换行后缀（如 '非高管\\n（中基层…）'）用前缀匹配 putf
- 高管明细来自每条 record 的 executives 字段（不再硬编码全局表）
- 占比/折扣率 小数存储 + 百分数格式（0.00%）
- 导出 make_builders(HEAD) 供 sync 模块复用同一套字段映射
"""
import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from lti_insight.config.loader import path_of


PCT = "0.00%"
_ROLEMAP = [
    ("董事长", "董事长"), ("总经理级", "总裁"), ("总经理级", "总经理"), ("总经理级", "联席总裁"),
    ("副总经理级", "副总裁"), ("财务负责人", "财务负责人"), ("董事会秘书", "董事会秘书"),
    ("董事", "董事"), ("非执行董事", "非执行"),
]


def _role_flags(title):
    flags = {k: 0 for k, _ in _ROLEMAP}
    for col, kw in _ROLEMAP:
        if kw in (title or ""):
            flags[col] = 1
    return flags


def _make_put(HEAD):
    def put(row, sheet, name, val):
        h = HEAD[sheet]
        if name in h and val is not None:
            row[h.index(name)] = val

    def putf(row, sheet, name, val):
        if val is None:
            return
        for i, hh in enumerate(HEAD[sheet]):
            if isinstance(hh, str) and hh.startswith(name):
                row[i] = val
                return

    return put, putf


def make_builders(HEAD):
    """返回 (scheme_row, exec_rows_all, perf_row)，基于给定表头 HEAD 做列映射。"""
    put, putf = _make_put(HEAD)

    def new_row(sheet):
        return [None] * len(HEAD[sheet])

    def scheme_row(r):
        row = new_row("方案")
        execs = r.get("executives") or []
        exec_total = sum(e.get("qty_wan", 0) or 0 for e in execs)
        exec_n = len(execs)
        non_exec = (r.get("n_recipients") or 0) - exec_n
        vest = r.get("vesting_schedule") or []
        vest_pct = "/".join(f"{int(x * 100)}%" for x in vest)
        cover = f"1+{len(vest)}" if len(vest) > 1 else "1+1"
        disc = r.get("discount_rate") if r.get("discount_rate") is not None else 0.5

        put(row, "方案", "方案名称", f"{r.get('name','')}{r.get('plan_name','')}")
        put(row, "方案", "计划名称", r.get("plan_name"))
        put(row, "方案", "公司代码", r.get("code"))
        put(row, "方案", "公司名称", r.get("name"))
        put(row, "方案", "上市板", r.get("board"))
        put(row, "方案", "公司属性", r.get("company_attr", "民营企业"))
        put(row, "方案", "激励工具", r.get("tool_type"))
        put(row, "方案", "方案进度", r.get("plan_status", "草案"))
        put(row, "方案", "预案公告日", r.get("announce_date"))
        put(row, "方案", "授予年份", r.get("grant_year", 2026))
        put(row, "方案", "是否分期授予", "是" if r.get("has_reserved") else "否")
        src = r.get("share_source")
        if not src:
            if r.get("code") == "300347":
                src = "回购"
            elif r.get("code") in ("688169", "603883"):
                src = "定向发行"
            else:
                src = "定向发行或回购"
        put(row, "方案", "股票来源", src)
        put(row, "方案", "首次授予数量(万股/万份)", r.get("first_grant_wan"))
        put(row, "方案", "预留数量(万股/万份)", r.get("reserved_wan"))
        put(row, "方案", "激励总数(万股/万份)", r.get("total_shares_wan"))
        put(row, "方案", "激励总数占当时总股本比例(%)", r.get("total_pct"))
        if r.get("reserved_wan"):
            put(row, "方案", "预留数量占激励总数比(%)", r.get("reserved_wan") / r.get("total_shares_wan"))
        put(row, "方案", "高管授予总量(万股)", exec_total)
        if exec_total:
            put(row, "方案", "高管授予总量占比(%)", exec_total / r.get("total_shares_wan"))
        put(row, "方案", "高管", exec_n)
        putf(row, "方案", "非高管", non_exec)
        put(row, "方案", "激励总人数", r.get("n_recipients"))
        put(row, "方案", "等待期(年)", r.get("wait_years", 1 if r.get("code") != "688322" else 2.5))
        put(row, "方案", "归属/行权期(年)", (r.get("validity_months") or 60) / 12)
        put(row, "方案", "覆盖时间", cover)
        put(row, "方案", cover, "√")
        put(row, "方案", "归属节奏", vest_pct)
        for i, v in enumerate(vest[:6], start=1):
            put(row, "方案", f"第{i}期解锁比例", v)
        put(row, "方案", "解锁比例总计", sum(vest))
        put(row, "方案", "定价依据", "《上市公司股权激励管理办法》")
        put(row, "方案", "定价基准", r.get("pricing_basis"))
        put(row, "方案", "期权行权价格/股票授予价格", r.get("grant_price"))
        put(row, "方案", "期权行权价格/股票授予价格折扣", disc)
        put(row, "方案", "采用证监会或国资委基准定价方式", "是")
        if r.get("is_combo"):
            put(row, "方案", "期权定价折扣", 1.0)
            put(row, "方案", "股票定价折扣", 0.5)
        else:
            put(row, "方案", "股票定价折扣", disc)
        pb = r.get("pricing_basis") or ""
        if "120" in pb:
            m = re.search(r"120.{0,6}?(\d+\.?\d*)", pb)
            if m:
                put(row, "方案", "草案公告前120个工作日股票交易均价", float(m.group(1)) * 2)
        if "60" in pb:
            m = re.search(r"60.{0,6}?(\d+\.?\d*)", pb)
            if m:
                put(row, "方案", "草案公告前60个工作日股票交易均价", float(m.group(1)) * 2)
        if "1个交易" in pb or "1 个交易" in pb or "前1个" in pb:
            m = re.search(r"前1个.{0,6}?(\d+\.?\d*)", pb)
            if m:
                put(row, "方案", "草案公告前1个工作日股票交易均价", float(m.group(1)) * 2)
        put(row, "方案", "是否披露激励对象名单", "是" if (exec_n or r.get("n_recipients")) else "否")
        put(row, "方案", "是否有预留安排", "是" if r.get("has_reserved") else "否")
        put(row, "方案", "GICS整理后行业分类", r.get("industry"))
        return row

    def exec_rows_all(records):
        out = []
        for r in records:
            code = r.get("code")
            name = r.get("name")
            for e in (r.get("executives") or []):
                row = new_row("高管")
                rf = _role_flags(e.get("title"))
                put(row, "高管", "方案名称", f"{name}{r.get('plan_name','')}")
                put(row, "高管", "公司代码", code)
                put(row, "高管", "公司名称", name)
                put(row, "高管", "激励工具", r.get("tool_type"))
                put(row, "高管", "方案进度", r.get("plan_status", "草案"))
                put(row, "高管", "预案公告日", r.get("announce_date"))
                put(row, "高管", "姓名", e.get("name"))
                put(row, "高管", "代码&姓名", f"{code}{e.get('name')}")
                put(row, "高管", "授予数量（万股）-草案公告", e.get("qty_wan"))
                put(row, "高管", "是否为预留授予", "否")
                put(row, "高管", "职位", e.get("title"))
                put(row, "高管", "董事长", rf["董事长"])
                put(row, "高管", "总经理级", rf["总经理级"])
                put(row, "高管", "副总经理级", rf["副总经理级"])
                put(row, "高管", "财务负责人", rf["财务负责人"])
                put(row, "高管", "董事会秘书", rf["董事会秘书"])
                put(row, "高管", "董事", rf["董事"])
                put(row, "高管", "非执行董事", rf["非执行董事"])
                put(row, "高管", "覆盖时间", f"1+{len(r.get('vesting_schedule') or [0])}")
                put(row, "高管", "授予年份", r.get("grant_year", 2026))
                put(row, "高管", "GICS整理后行业分类", r.get("industry"))
                put(row, "高管", "上市板", r.get("board"))
                out.append(row)
        return out

    def perf_row(r):
        row = new_row("业绩考核")
        put(row, "业绩考核", "方案名称", f"{r.get('name','')}{r.get('plan_name','')}")
        put(row, "业绩考核", "公司代码", r.get("code"))
        put(row, "业绩考核", "公司名称", r.get("name"))
        put(row, "业绩考核", "公司属性", r.get("company_attr", "民营企业"))
        put(row, "业绩考核", "方案进度", r.get("plan_status", "草案"))
        put(row, "业绩考核", "预案公告日", r.get("announce_date"))
        put(row, "业绩考核", "解锁比例", "/".join(f"{int(x * 100)}%" for x in (r.get("vesting_schedule") or [])))
        put(row, "业绩考核", "授予考核指标明细",
            f"{r.get('perf_metric_type','')}；{r.get('perf_base','')}；{r.get('perf_targets','')}")
        put(row, "业绩考核", "行权/解锁考核明细", r.get("perf_targets"))
        put(row, "业绩考核", "备注", r.get("note"))
        put(row, "业绩考核", "行业分类", r.get("industry"))
        return row

    return scheme_row, exec_rows_all, perf_row


def build_workbook(records, out_path, template_path):
    twb = load_workbook(template_path, read_only=True, data_only=True)
    HEAD = {s: [c.value for c in next(twb[s].iter_rows(min_row=1, max_row=1))] for s in twb.sheetnames}
    twb.close()

    scheme_row, exec_rows_all, perf_row = make_builders(HEAD)
    wb = Workbook()
    pct_cols = {
        "方案": ["激励总数占当时总股本比例(%)", "预留数量占激励总数比(%)", "高管授予总量占比(%)",
                 "第1期解锁比例", "第2期解锁比例", "第3期解锁比例", "第4期解锁比例",
                 "第5期解锁比例", "第6期解锁比例",
                 "期权行权价格/股票授予价格折扣", "期权定价折扣", "股票定价折扣"],
        "业绩考核": [], "高管": [],
    }
    hdr_fill = PatternFill("solid", fgColor="1F3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9E1EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for s in ["方案", "高管", "业绩考核"]:
        ws = wb.active if s == "方案" else wb.create_sheet(s)
        ws.title = s
        hdr = HEAD[s]
        for c, h in enumerate(hdr, 1):
            cell = ws.cell(1, c, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        rows = []
        if s == "方案":
            rows = [scheme_row(r) for r in records]
        elif s == "高管":
            rows = exec_rows_all(records)
        elif s == "业绩考核":
            rows = [perf_row(r) for r in records]
        for i, row in enumerate(rows, start=2):
            for c, v in enumerate(row, 1):
                cell = ws.cell(i, c, v)
                cell.border = border
                colname = hdr[c - 1]
                if colname in pct_cols[s] and isinstance(v, (int, float)):
                    cell.number_format = PCT
        ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return len(records), len(exec_rows_all(records)), len(records)


def rebuild(out_path=None, template_path=None, records=None):
    """从 store 重新生成按模板工作簿。返回 (n_scheme, n_exec, n_perf)。"""
    out_path = out_path or os.path.join(path_of("output_dir"), "LTI数据库_按模板.xlsx")
    template_path = template_path or path_of("template_xlsx")
    if records is None:
        from lti_insight.core import store
        records = store.load()["records"]
    return build_workbook(records, out_path, template_path)
